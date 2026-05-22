# ============================================================
# DTM Survey Alignment Tool — exporter.py
# Python port of export_alignment_excel() from R (v10)
# Dependencies: openpyxl, pandas
# ============================================================

import io
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from typing import Dict, List, Optional, Tuple
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────
# Colour palette  (mirrors R .COLOURS list)
# ────────────────────────────────────────────────────────────

_C = {
    "dark_blue":  "1F3864",
    "mid_blue":   "2E5797",
    "light_blue": "D6E4F7",
    "grey_row":   "F2F2F2",
    "white":      "FFFFFF",
}

_STATUS_COLOURS: Dict[str, Tuple[str, str]] = {
    "LikelyAligned":       ("C6EFCE", "375623"),
    "NeedsVerification":   ("FFEB9C", "9C6500"),
    "FPReview":            ("FFC7CE", "9C0006"),
    "DiscouragedQuestion": ("FCE4D6", "833C00"),
    "MissingCoreQuestion": ("E2D9F3", "4B2E83"),
    "DoesNotNeed2Align":   ("F5F5F5", "666666"),
}

_BEGIN_GROUP_BG = "4066B8"
_BEGIN_GROUP_FG = "FFFFFF"


def _status_colours(status: str) -> Tuple[str, str]:
    """Return (fill_hex, font_hex) for a given AlignmentStatus value."""
    if isinstance(status, str) and status.startswith("Missing"):
        return _STATUS_COLOURS["MissingCoreQuestion"]
    return _STATUS_COLOURS.get(status, ("FFFFFF", "000000"))


# ────────────────────────────────────────────────────────────
# Style factory helpers
# ────────────────────────────────────────────────────────────

def _fill(hex_col: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_col)


def _font(hex_col: str, bold: bool = False, size: int = 10) -> Font:
    return Font(name="Arial", size=size, color=hex_col, bold=bold)


def _border() -> Border:
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical="top", wrap_text=wrap)


def _cell_val(val):
    """Convert pandas NA types to None so openpyxl writes blank cells."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


# ────────────────────────────────────────────────────────────
# Column-width lookup map
# ────────────────────────────────────────────────────────────

_COL_WIDTHS: Dict[str, float] = {
    "SN": 5,
    "type": 18,
    "name": 30,
    "label": 50,
    "label::English (en)": 50,
    "label::Portuguese (pt)": 40,
    "hint::English (en)": 30,
    "hint::Portuguese (pt)": 30,
    "appearance": 14,
    "required": 10,
    "relevant": 25,
    "constraint": 20,
    "calculation": 20,
    "choice_filter": 15,
    "Matching": 38,
    "AlignmentStatus": 22,
    "ComponentMatch": 18,
}

_DK_COL_WIDTHS: Dict[str, float] = {
    "FieldName": 30,
    "FieldUniqueId": 18,
    "QuestionText(en)": 50,
    "QuestionAnswerType": 22,
    "QuestionComponent": 55,
}


def _col_width(col_name: str, df2_name: str = "Datakit") -> float:
    """Return an appropriate column width for a given column name."""
    if col_name in _COL_WIDTHS:
        return _COL_WIDTHS[col_name]
    # Strip df2 prefix to look up datakit columns
    prefix = f"{df2_name}_"
    if col_name.startswith(prefix):
        base = col_name[len(prefix):]
        if base in _DK_COL_WIDTHS:
            return _DK_COL_WIDTHS[base]
    return 20.0


# ────────────────────────────────────────────────────────────
# Main export function
# ────────────────────────────────────────────────────────────

def export_alignment_excel(
    result_df: pd.DataFrame,
    df1_name: str = "survey",
    df2_name: str = "Datakit",
    df1_key_col: str = "name",
    df1_text_col: str = "label",
    formcomponents: str = "Baseline Sub-Area Assessment",
) -> io.BytesIO:
    """
    Build a styled three-sheet Excel workbook from the match_surveys() output
    and return it as an in-memory BytesIO (ready for st.download_button).

    Sheets:
        Summary                      — dashboard + legend
        Survey_Alignment_Diagnostics — all rows, colour-coded
        Missing_BSA_Core             — only Missing* rows
    """

    wb = Workbook()
    wb.remove(wb.active)   # drop the default blank sheet

    # ── Classify columns ────────────────────────────────────
    diag_cols = [c for c in ["Matching", "AlignmentStatus", "ComponentMatch"]
                 if c in result_df.columns]
    misc_cols = [".RowSource"]

    # Datakit output columns (those prefixed with df2_name)
    dk_cols = [
        c for c in result_df.columns
        if c.startswith(f"{df2_name}_")
    ]

    # Original survey columns = everything else
    appended = set(diag_cols + dk_cols + misc_cols)
    survey_cols = [c for c in result_df.columns if c not in appended]

    # ── Stats for Summary sheet ─────────────────────────────
    is_df1 = result_df[".RowSource"] == "df1"
    df1_rows = result_df[is_df1]
    n_df1 = len(df1_rows)

    status_vc = result_df["AlignmentStatus"].value_counts()
    n_likely = int(status_vc.get("LikelyAligned", 0))
    n_verif = int(status_vc.get("NeedsVerification", 0))
    n_fpr = int(status_vc.get("FPReview", 0))
    n_disc = int(status_vc.get("DiscouragedQuestion", 0))
    n_miss_core = sum(
        v for k, v in status_vc.items()
        if isinstance(k, str) and k.startswith("Missing")
    )
    n_missing_appended = int((~is_df1).sum())

    comp_col_present = "ComponentMatch" in result_df.columns
    n_comp_yes = int((result_df["ComponentMatch"] ==
                     "Yes").sum()) if comp_col_present else 0
    n_comp_no = int((result_df["ComponentMatch"] ==
                    "No").sum()) if comp_col_present else 0

    def _pct(x: int, tot: int = n_df1) -> str:
        if tot == 0:
            return "0.0%"
        return f"{x / tot * 100:.1f}%"

    # ── SHEET 1: Summary ────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Row 1 — title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"Survey Alignment Diagnostic — {formcomponents}"
    c.fill = _fill(_C["dark_blue"])
    c.font = _font("FFFFFF", bold=True, size=14)
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 30

    # Row 2 — subtitle
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"{df1_name} vs. {df2_name} Global Data Kit"
    c.fill = _fill(_C["light_blue"])
    c.font = Font(name="Arial", size=10, color="444444", italic=True)
    c.alignment = _align(h="center")
    ws.row_dimensions[2].height = 18

    # Row 4 — column headers
    for col, label in enumerate(["Metric", "Count", "% of Survey"], 1):
        c = ws.cell(row=4, column=col, value=label)
        c.fill = _fill(_C["dark_blue"])
        c.font = _font("FFFFFF", bold=True)
        c.alignment = _align(h="center")
        c.border = _border()
    ws.row_dimensions[4].height = 18

    # Rows 5+ — stats data
    stats_rows = [
        ("Total Survey Questions",           n_df1,      "100%"),
        ("",                                 None,       ""),
        ("Likely Aligned  (≥85% match)",     n_likely,   _pct(n_likely)),
        ("Needs Verification (fuzzy match)", n_verif,    _pct(n_verif)),
        ("FP Review (no match)",             n_fpr,      _pct(n_fpr)),
        ("Discouraged Question",             n_disc,     _pct(n_disc)),
        (f"Missing {formcomponents} Core",   n_miss_core, _pct(n_miss_core)),
        ("",                                 None,       ""),
        ("Component Match — Yes",            n_comp_yes, _pct(n_comp_yes)),
        ("Component Match — No",             n_comp_no,  _pct(n_comp_no)),
    ]

    _STAT_KW: Dict[str, str] = {
        "likely aligned":    "LikelyAligned",
        "needs verification": "NeedsVerification",
        "fp review":         "FPReview",
        "discouraged":       "DiscouragedQuestion",
        "missing":           "MissingCoreQuestion",
    }

    for i, (metric, count, pct_val) in enumerate(stats_rows):
        row = 5 + i
        ws.row_dimensions[row].height = 18
        if not metric:
            continue

        # Detect which status colour to apply
        matched_status = None
        for kw, st in _STAT_KW.items():
            if kw in metric.lower():
                matched_status = st
                break

        if matched_status:
            bg, fg = _status_colours(matched_status)
            bold = True
        else:
            bg = _C["grey_row"] if i % 2 == 0 else _C["white"]
            fg = "000000"
            bold = False

        for col, val in enumerate([metric, count, pct_val], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = _fill(bg)
            c.font = _font(fg, bold=bold)
            c.alignment = _align()
            c.border = _border()

    # Legend section
    leg_start = 5 + len(stats_rows) + 2
    ws.merge_cells(f"A{leg_start}:F{leg_start}")
    c = ws.cell(row=leg_start, column=1, value="ALIGNMENT STATUS LEGEND")
    c.fill = _fill(_C["mid_blue"])
    c.font = _font("FFFFFF", bold=True)
    c.alignment = _align(h="center")
    ws.row_dimensions[leg_start].height = 18

    legend_items = [
        ("LikelyAligned",
         "Match score ≥85% — question confidently maps to Datakit entry"),
        ("NeedsVerification",
         "Fuzzy match found — manual confirmation recommended"),
        ("FPReview",
         "No match found — focal point to review / map manually"),
        ("DiscouragedQuestion",
         "Matched Datakit entry is flagged Discouraged for this component"),
        ("MissingCoreQuestion",
         f"Core question in Datakit not present in {df1_name} survey"),
        ("DoesNotNeed2Align",
         "Structural element (begin_group, calculate, note, etc.) — skipped"),
    ]

    for k, (status, desc) in enumerate(legend_items):
        r = leg_start + 1 + k
        bg, fg = _status_colours(status)
        c1 = ws.cell(row=r, column=1, value=status)
        c1.fill = _fill(bg)
        c1.font = _font(fg, bold=True)
        c1.border = _border()
        c1.alignment = _align()
        ws.merge_cells(f"B{r}:F{r}")
        c2 = ws.cell(row=r, column=2, value=desc)
        c2.fill = _fill(bg)
        c2.font = _font(fg)
        c2.border = _border()
        c2.alignment = _align()
        ws.row_dimensions[r].height = 18

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    for col_letter in ["D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 4

    # ── Shared data-sheet writer ─────────────────────────────

    def _write_data_sheet(sheet_name: str, df: pd.DataFrame, freeze_col: int = 1):
        """
        Write a colour-coded data sheet.
        Column order: SN | survey cols | diagnostic cols | datakit cols
        """
        ws_d = wb.create_sheet(sheet_name)
        ws_d.sheet_view.showGridLines = False

        # Build ordered column list
        diag_present = [c for c in diag_cols if c in df.columns]
        dk_present = [c for c in dk_cols if c in df.columns]
        surv_present = [c for c in survey_cols if c in df.columns]
        ordered_cols = surv_present + diag_present + dk_present

        # Build output df (SN prepended)
        out = df[ordered_cols].copy().reset_index(drop=True)
        sn_col = "SN1" if "SN" in out.columns else "SN"
        out.insert(0, sn_col, range(1, len(out) + 1))

        headers = list(out.columns)
        n_cols = len(headers)
        n_rows = len(out)

        # Header row
        for col_idx, hdr in enumerate(headers, 1):
            c = ws_d.cell(row=1, column=col_idx, value=hdr)
            c.fill = _fill(_C["dark_blue"])
            c.font = _font("FFFFFF", bold=True)
            c.alignment = _align(h="center")
            c.border = _border()
        ws_d.row_dimensions[1].height = 22

        # Identify special column positions (1-based)
        type_col_pos = (headers.index("type") +
                        1) if "type" in headers else None
        status_col_pos = (headers.index("AlignmentStatus") + 1) \
            if "AlignmentStatus" in headers else None

        # Data rows
        for row_idx in range(n_rows):
            r = row_idx + 2
            ws_d.row_dimensions[r].height = 14

            # Detect row type for begin_group banding
            row_type = ""
            if type_col_pos:
                raw_type = out.iat[row_idx, type_col_pos - 1]
                row_type = _safe_str(raw_type).lower()

            is_begin_group = (row_type == "begin_group")

            # Detect alignment status
            status = None
            if status_col_pos:
                s_val = out.iat[row_idx, status_col_pos - 1]
                if s_val and not (isinstance(s_val, float) and np.isnan(s_val)):
                    status = str(s_val)

            for col_idx, col_name in enumerate(headers, 1):
                raw = out.iat[row_idx, col_idx - 1]
                val = _cell_val(raw)
                c = ws_d.cell(row=r, column=col_idx, value=val)
                c.border = _border()
                c.alignment = _align()

                if is_begin_group:
                    c.fill = _fill(_BEGIN_GROUP_BG)
                    c.font = _font(_BEGIN_GROUP_FG, bold=True)
                elif col_name == "AlignmentStatus" and status:
                    bg, fg = _status_colours(status)
                    c.fill = _fill(bg)
                    c.font = _font(fg, bold=True)
                else:
                    row_bg = _C["grey_row"] if row_idx % 2 == 0 else _C["white"]
                    c.fill = _fill(row_bg)
                    c.font = _font("000000")

        # Auto-filter on header row
        ws_d.auto_filter.ref = (
            f"A1:{get_column_letter(n_cols)}1"
        )

        # Freeze panes: freeze row 1 + first `freeze_col` columns
        ws_d.freeze_panes = ws_d.cell(row=2, column=freeze_col + 1)

        # Column widths
        for col_idx, col_name in enumerate(headers, 1):
            ws_d.column_dimensions[get_column_letter(col_idx)].width = (
                _col_width(col_name, df2_name)
            )

    # ── SHEET 2: Survey_Alignment_Diagnostics ───────────────
    _write_data_sheet("Survey_Alignment_Diagnostics", result_df, freeze_col=2)

    # ── SHEET 3: Missing_BSA_Core ────────────────────────────
    missing_mask = result_df["AlignmentStatus"].str.startswith(
        "Missing", na=False)
    if missing_mask.any():
        _write_data_sheet("Missing_BSA_Core",
                          result_df[missing_mask], freeze_col=1)

    # ── Serialise to BytesIO ─────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _safe_str(val) -> str:
    """Local safe-string converter (avoids circular import from matcher)."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val).strip()
